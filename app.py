import os
import io
from flask import Flask, render_template, request, redirect, url_for, session, flash, jsonify, send_file
from flask_mysqldb import MySQL
from werkzeug.security import generate_password_hash, check_password_hash
import json
from decimal import Decimal
from datetime import datetime, timedelta
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
app.secret_key = 'sppg-inventaris-mbg-2024'

# Konfigurasi database MySQL
app.config['MYSQL_HOST'] = 'localhost'
app.config['MYSQL_USER'] = 'root'
app.config['MYSQL_PASSWORD'] = ''
app.config['MYSQL_DB'] = 'inventaris_sppg'
app.config['MYSQL_CURSORCLASS'] = 'DictCursor'

mysql = MySQL(app)

# Custom Filter untuk format angka (ribuan)
@app.template_filter('format_number')
def format_number(value):
    """Format angka dengan pemisah ribuan"""
    if value is None:
        return "0"
    try:
        # Konversi ke float dulu
        if isinstance(value, Decimal):
            value = float(value)
        elif isinstance(value, str):
            try:
                value = float(value)
            except ValueError:
                return str(value)
        
        # Format dengan pemisah ribuan
        return "{:,.0f}".format(value).replace(",", ".")
    except (ValueError, TypeError, AttributeError) as e:
        print(f"Error formatting number {value}: {e}")
        return str(value) if value is not None else "0"

# Fungsi helper untuk format currency
@app.template_filter('format_currency')
def format_currency(value):
    """Format mata uang Rupiah"""
    if value is None:
        return "Rp 0"
    try:
        if isinstance(value, Decimal):
            value = float(value)
        return "Rp {:,.0f}".format(value).replace(",", ".")
    except (ValueError, TypeError):
        try:
            return "Rp {:,.0f}".format(float(value)).replace(",", ".")
        except:
            return "Rp 0"

# Filter untuk konversi ke float
@app.template_filter('to_float')
def to_float_filter(value):
    """Konversi ke float untuk operasi matematika di template"""
    try:
        if value is None:
            return 0.0
        if isinstance(value, Decimal):
            return float(value)
        return float(value)
    except (ValueError, TypeError):
        return 0.0

# Fungsi untuk memeriksa login
def is_logged_in():
    return 'logged_in' in session

# Middleware untuk memeriksa autentikasi
@app.before_request
def require_login():
    allowed_routes = ['login', 'static']
    if request.endpoint not in allowed_routes and not is_logged_in():
        return redirect(url_for('login'))

# Fungsi untuk memeriksa role
def check_role(required_roles):
    if 'role' not in session:
        return False
    return session['role'] in required_roles

# Fungsi helper untuk menghitung total
def calculate_total(jumlah, harga_satuan):
    # Konversi ke Decimal jika belum
    if isinstance(jumlah, float):
        jumlah = Decimal(str(jumlah))
    if isinstance(harga_satuan, float):
        harga_satuan = Decimal(str(harga_satuan))
    
    # Sekarang bisa dikalikan
    return jumlah * harga_satuan

# Halaman Login
@app.route('/', methods=['GET', 'POST'])
@app.route('/login', methods=['GET', 'POST'])
def login():
    if is_logged_in():
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        cur = mysql.connection.cursor()
        cur.execute("SELECT * FROM users WHERE username = %s", [username])
        user = cur.fetchone()
        cur.close()
        
        if user:
            # Untuk demo, password adalah 'admin123' (tanpa hash)
            if password == user['password']:
                session['logged_in'] = True
                session['user_id'] = user['id']
                session['username'] = user['username']
                session['nama_lengkap'] = user['nama_lengkap']
                session['role'] = user['role']
                flash('Login berhasil!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Password salah!', 'danger')
        else:
            flash('Username tidak ditemukan!', 'danger')
    
    return render_template('login.html')

# Logout
@app.route('/logout')
def logout():
    session.clear()
    flash('Anda telah logout', 'info')
    return redirect(url_for('login'))

# Dashboard
@app.route('/dashboard')
def dashboard():
    cur = mysql.connection.cursor()
    
    # Total bahan
    cur.execute("SELECT COUNT(*) as total FROM bahan WHERE status = 'aktif'")
    total_bahan = cur.fetchone()['total']
    
    # Total stok
    cur.execute("SELECT SUM(jumlah) as total FROM stok")
    total_stok_result = cur.fetchone()['total']
    total_stok = float(total_stok_result) if total_stok_result else 0.0
    
    # Total penerimaan bulan ini
    bulan_ini = datetime.now().strftime('%Y-%m')
    cur.execute("SELECT SUM(jumlah) as total FROM penerimaan WHERE DATE_FORMAT(tanggal, '%%Y-%%m') = %s AND status = 'disetujui'", [bulan_ini])
    total_penerimaan_result = cur.fetchone()['total']
    total_penerimaan = float(total_penerimaan_result) if total_penerimaan_result else 0.0
    
    # Total pengeluaran bulan ini
    cur.execute("SELECT SUM(jumlah) as total FROM pengeluaran WHERE DATE_FORMAT(tanggal, '%%Y-%%m') = %s AND status != 'draft'", [bulan_ini])
    total_pengeluaran_result = cur.fetchone()['total']
    total_pengeluaran = float(total_pengeluaran_result) if total_pengeluaran_result else 0.0
    
    # Bahan hampir habis (stok < stok_minimum)
    cur.execute("""
        SELECT b.nama_bahan, s.jumlah, b.stok_minimum, sat.nama_satuan
        FROM stok s
        JOIN bahan b ON s.bahan_id = b.id
        JOIN satuan sat ON s.satuan_id = sat.id
        WHERE s.jumlah <= b.stok_minimum
        ORDER BY s.jumlah ASC
        LIMIT 5
    """)
    bahan_hampir_habis = cur.fetchall()
    
    # Konversi Decimal ke float
    for item in bahan_hampir_habis:
        for key in ['jumlah', 'stok_minimum']:
            if isinstance(item.get(key), Decimal):
                item[key] = float(item[key])
    
    # Bahan mendekati kadaluarsa (30 hari ke depan)
    hari_ini = datetime.now().date()
    tiga_puluh_hari = (datetime.now() + timedelta(days=30)).date()
    cur.execute("""
        SELECT p.no_penerimaan, b.nama_bahan, p.jumlah, s.nama_satuan, p.tanggal_kadaluarsa,
               DATEDIFF(p.tanggal_kadaluarsa, CURDATE()) as hari_menuju_kadaluarsa
        FROM penerimaan p
        JOIN bahan b ON p.bahan_id = b.id
        JOIN satuan s ON p.satuan_id = s.id
        WHERE p.tanggal_kadaluarsa BETWEEN %s AND %s
        AND p.status = 'disetujui'
        ORDER BY p.tanggal_kadaluarsa ASC
        LIMIT 5
    """, [hari_ini, tiga_puluh_hari])
    bahan_mendekati_kadaluarsa = cur.fetchall()
    
    # Data untuk grafik stok per kategori
    cur.execute("""
        SELECT kb.nama_kategori, SUM(s.jumlah) as total_stok
        FROM stok s
        JOIN bahan b ON s.bahan_id = b.id
        JOIN kategori_bahan kb ON b.kategori_id = kb.id
        GROUP BY kb.id, kb.nama_kategori
        ORDER BY total_stok DESC
    """)
    stok_per_kategori = cur.fetchall()
    
    # Data untuk grafik distribusi per tujuan
    cur.execute("""
        SELECT jenis_tujuan, COUNT(*) as jumlah
        FROM pengeluaran
        WHERE status != 'draft'
        GROUP BY jenis_tujuan
        ORDER BY jumlah DESC
    """)
    distribusi_per_tujuan = cur.fetchall()
    
    # Monitoring kualitas terbaru
    cur.execute("""
        SELECT mk.tanggal_check, b.nama_bahan, mk.kondisi_fisik, mk.status_kadaluarsa, mk.petugas
        FROM monitoring_kualitas mk
        JOIN bahan b ON mk.bahan_id = b.id
        ORDER BY mk.tanggal_check DESC
        LIMIT 5
    """)
    monitoring_terbaru = cur.fetchall()
    
    cur.close()
    
    return render_template('dashboard.html', 
                         total_bahan=total_bahan,
                         total_stok=total_stok,
                         total_penerimaan=total_penerimaan,
                         total_pengeluaran=total_pengeluaran,
                         bahan_hampir_habis=bahan_hampir_habis,
                         bahan_mendekati_kadaluarsa=bahan_mendekati_kadaluarsa,
                         stok_per_kategori=stok_per_kategori,
                         distribusi_per_tujuan=distribusi_per_tujuan,
                         monitoring_terbaru=monitoring_terbaru)

# API untuk data chart
@app.route('/api/stok-per-kategori')
def api_stok_per_kategori():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT kb.nama_kategori, SUM(s.jumlah) as total_stok
        FROM stok s
        JOIN bahan b ON s.bahan_id = b.id
        JOIN kategori_bahan kb ON b.kategori_id = kb.id
        GROUP BY kb.id, kb.nama_kategori
        ORDER BY total_stok DESC
    """)
    data = cur.fetchall()
    cur.close()
    
    labels = [item['nama_kategori'] for item in data]
    values = [float(item['total_stok']) if item['total_stok'] else 0.0 for item in data]
    
    return jsonify({
        'labels': labels,
        'values': values
    })

@app.route('/api/penerimaan-per-bulan')
def api_penerimaan_per_bulan():
    cur = mysql.connection.cursor()
    cur.execute("""
        SELECT MONTH(tanggal) as bulan, SUM(jumlah) as total 
        FROM penerimaan 
        WHERE YEAR(tanggal) = YEAR(CURDATE()) AND status = 'disetujui'
        GROUP BY MONTH(tanggal) 
        ORDER BY bulan
    """)
    data = cur.fetchall()
    cur.close()
    
    labels = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Agu', 'Sep', 'Okt', 'Nov', 'Des']
    values = [0] * 12
    
    for item in data:
        if 1 <= item['bulan'] <= 12:
            values[item['bulan'] - 1] = float(item['total']) if item['total'] else 0.0
    
    return jsonify({
        'labels': labels,
        'values': values
    })

# Master Data - Bahan
@app.route('/master-bahan')
def master_bahan():
    if not check_role(['admin', 'gudang']):
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Pencarian
    keyword = request.args.get('keyword', '')
    filter_kategori = request.args.get('kategori', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT b.*, kb.nama_kategori, s.nama_satuan, COALESCE(st.jumlah, 0) as stok_sekarang
        FROM bahan b
        LEFT JOIN kategori_bahan kb ON b.kategori_id = kb.id
        LEFT JOIN satuan s ON b.satuan_id = s.id
        LEFT JOIN stok st ON b.id = st.bahan_id
        WHERE b.status = 'aktif'
    """
    params = []
    
    if keyword:
        query += " AND (b.nama_bahan LIKE %s OR b.kode_bahan LIKE %s)"
        params.extend([f"%{keyword}%", f"%{keyword}%"])
    
    if filter_kategori:
        query += " AND b.kategori_id = %s"
        params.append(filter_kategori)
    
    query += " ORDER BY b.nama_bahan"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Konversi Decimal ke float untuk field tertentu dan hitung statistik
    stok_kritis = 0
    total_stok = 0
    
    for item in data:
        for key in ['stok_sekarang', 'stok_minimum', 'stok_maksimum', 'berat_per_unit', 'kalori_per_unit', 'protein_per_unit']:
            if item.get(key) is not None and isinstance(item[key], Decimal):
                item[key] = float(item[key])
        
        # Hitung stok kritis dan total stok
        stok_sekarang = item.get('stok_sekarang', 0) or 0
        stok_minimum = item.get('stok_minimum', 0) or 0
        
        if isinstance(stok_sekarang, (int, float)) and isinstance(stok_minimum, (int, float)):
            total_stok += stok_sekarang
            if stok_sekarang <= stok_minimum:
                stok_kritis += 1
    
    # Ambil list kategori unik untuk filter
    cur.execute("SELECT * FROM kategori_bahan ORDER BY nama_kategori")
    kategori_list = cur.fetchall()
    
    cur.close()
    
    return render_template('master_bahan.html', 
                         bahan=data, 
                         keyword=keyword,
                         filter_kategori=filter_kategori,
                         kategori_list=kategori_list,
                         stok_kritis=stok_kritis,
                         total_stok=total_stok)

# Tambah Bahan
@app.route('/tambah-bahan', methods=['GET', 'POST'])
def tambah_bahan():
    if not check_role(['admin', 'gudang']):
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    cur = mysql.connection.cursor()
    
    if request.method == 'POST':
        kode_bahan = request.form['kode_bahan']
        nama_bahan = request.form['nama_bahan']
        kategori_id = request.form['kategori_id']
        satuan_id = request.form['satuan_id']
        stok_minimum = float(request.form.get('stok_minimum', 0) or 0)
        stok_maksimum = float(request.form.get('stok_maksimum', 0) or 0)
        berat_per_unit = float(request.form.get('berat_per_unit', 1) or 1)
        kalori_per_unit = float(request.form.get('kalori_per_unit', 0) or 0)
        protein_per_unit = float(request.form.get('protein_per_unit', 0) or 0)
        keterangan = request.form.get('keterangan', '')
        
        try:
            cur.execute("""
                INSERT INTO bahan (kode_bahan, nama_bahan, kategori_id, satuan_id, 
                                 stok_minimum, stok_maksimum, berat_per_unit,
                                 kalori_per_unit, protein_per_unit, keterangan)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (kode_bahan, nama_bahan, kategori_id, satuan_id, 
                  stok_minimum, stok_maksimum, berat_per_unit,
                  kalori_per_unit, protein_per_unit, keterangan))
            
            bahan_id = cur.lastrowid
            
            # Buat stok awal 0
            cur.execute("INSERT INTO stok (bahan_id, jumlah, satuan_id) VALUES (%s, 0, %s)", 
                       (bahan_id, satuan_id))
            
            mysql.connection.commit()
            flash('Data bahan berhasil ditambahkan!', 'success')
            return redirect(url_for('master_bahan'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    # Ambil data untuk dropdown
    cur.execute("SELECT * FROM kategori_bahan ORDER BY nama_kategori")
    kategori_list = cur.fetchall()
    
    cur.execute("SELECT * FROM satuan ORDER BY nama_satuan")
    satuan_list = cur.fetchall()
    
    cur.close()
    
    return render_template('tambah_bahan.html', 
                         kategori_list=kategori_list,
                         satuan_list=satuan_list)

# Penerimaan Barang
@app.route('/penerimaan')
def penerimaan():
    if not check_role(['admin', 'gudang']):
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Filter
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT p.*, b.nama_bahan, b.kode_bahan, s.nama_satuan,
               COALESCE(p.total_harga, p.jumlah * p.harga_satuan) as total_harga
        FROM penerimaan p
        JOIN bahan b ON p.bahan_id = b.id
        JOIN satuan s ON p.satuan_id = s.id
        WHERE 1=1
    """
    params = []
    
    if start_date:
        query += " AND p.tanggal >= %s"
        params.append(start_date)
    
    if end_date:
        query += " AND p.tanggal <= %s"
        params.append(end_date)
    
    if status:
        query += " AND p.status = %s"
        params.append(status)
    
    query += " ORDER BY p.tanggal DESC"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Konversi Decimal ke float untuk field tertentu
    for item in data:
        for key in ['jumlah', 'harga_satuan', 'total_harga']:
            if item.get(key) is not None and isinstance(item[key], Decimal):
                item[key] = float(item[key])
    
    cur.close()
    
    # Ambil tanggal hari ini untuk template
    today = datetime.now().date()
    
    return render_template('penerimaan.html', 
                         penerimaan=data,
                         start_date=start_date,
                         end_date=end_date,
                         status=status,
                         today=today)

# Tambah Penerimaan
@app.route('/tambah-penerimaan', methods=['GET', 'POST'])
def tambah_penerimaan():
    if not check_role(['admin', 'gudang']):
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    cur = mysql.connection.cursor()
    
    if request.method == 'POST':
        no_penerimaan = request.form['no_penerimaan']
        tanggal = request.form['tanggal']
        bahan_id = request.form['bahan_id']
        jumlah = float(request.form['jumlah'])
        satuan_id = request.form['satuan_id']
        harga_satuan = float(request.form.get('harga_satuan', 0) or 0)
        total_harga = jumlah * harga_satuan  # Hitung total
        supplier = request.form.get('supplier', '')
        no_batch = request.form.get('no_batch', '')
        tanggal_produksi = request.form.get('tanggal_produksi', None)
        tanggal_kadaluarsa = request.form.get('tanggal_kadaluarsa', None)
        kondisi = request.form.get('kondisi', 'baik')
        penerima = request.form.get('penerima', '')
        catatan = request.form.get('catatan', '')
        
        try:
            # Simpan penerimaan dengan total_harga
            cur.execute("""
                INSERT INTO penerimaan 
                (no_penerimaan, tanggal, bahan_id, jumlah, satuan_id, harga_satuan,
                 total_harga, supplier, no_batch, tanggal_produksi, tanggal_kadaluarsa,
                 kondisi, penerima, catatan, status)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'disetujui')
            """, (no_penerimaan, tanggal, bahan_id, jumlah, satuan_id, harga_satuan,
                  total_harga, supplier, no_batch, tanggal_produksi, tanggal_kadaluarsa,
                  kondisi, penerima, catatan))
            
            # Update stok
            cur.execute("""
                UPDATE stok 
                SET jumlah = jumlah + %s,
                    last_update = CURRENT_TIMESTAMP
                WHERE bahan_id = %s
            """, (jumlah, bahan_id))
            
            mysql.connection.commit()
            flash('Penerimaan berhasil dicatat dan stok diperbarui!', 'success')
            return redirect(url_for('penerimaan'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    # Ambil data untuk dropdown
    cur.execute("SELECT * FROM bahan WHERE status = 'aktif' ORDER BY nama_bahan")
    bahan_list = cur.fetchall()
    
    cur.execute("SELECT * FROM satuan ORDER BY nama_satuan")
    satuan_list = cur.fetchall()
    
    # Generate nomor penerimaan otomatis
    today = datetime.now().strftime('%Y%m%d')
    cur.execute("SELECT COUNT(*) as count FROM penerimaan WHERE DATE(created_at) = CURDATE()")
    count = cur.fetchone()['count']
    no_penerimaan_otomatis = f"TRM-{today}-{count+1:03d}"
    
    cur.close()
    
    return render_template('tambah_penerimaan.html', 
                         bahan_list=bahan_list,
                         satuan_list=satuan_list,
                         no_penerimaan_otomatis=no_penerimaan_otomatis)

# Pengeluaran Barang
@app.route('/pengeluaran')
def pengeluaran():
    if not check_role(['admin', 'gudang', 'distribusi']):
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Filter
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    status = request.args.get('status', '')
    jenis_tujuan = request.args.get('jenis_tujuan', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT p.*, b.nama_bahan, b.kode_bahan, s.nama_satuan
        FROM pengeluaran p
        JOIN bahan b ON p.bahan_id = b.id
        JOIN satuan s ON p.satuan_id = s.id
        WHERE 1=1
    """
    params = []
    
    if start_date:
        query += " AND p.tanggal >= %s"
        params.append(start_date)
    
    if end_date:
        query += " AND p.tanggal <= %s"
        params.append(end_date)
    
    if status:
        query += " AND p.status = %s"
        params.append(status)
    
    if jenis_tujuan:
        query += " AND p.jenis_tujuan = %s"
        params.append(jenis_tujuan)
    
    query += " ORDER BY p.tanggal DESC"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Konversi Decimal ke float untuk field tertentu
    for item in data:
        for key in ['jumlah']:
            if item.get(key) is not None and isinstance(item[key], Decimal):
                item[key] = float(item[key])
    
    cur.close()
    
    return render_template('pengeluaran.html', 
                         pengeluaran=data,
                         start_date=start_date,
                         end_date=end_date,
                         status=status,
                         jenis_tujuan=jenis_tujuan)

# Tambah Pengeluaran
@app.route('/tambah-pengeluaran', methods=['GET', 'POST'])
def tambah_pengeluaran():
    if not check_role(['admin', 'gudang', 'distribusi']):
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    cur = mysql.connection.cursor()
    
    if request.method == 'POST':
        no_pengeluaran = request.form['no_pengeluaran']
        tanggal = request.form['tanggal']
        bahan_id = request.form['bahan_id']
        jumlah = float(request.form['jumlah'])
        satuan_id = request.form['satuan_id']
        tujuan = request.form['tujuan']
        jenis_tujuan = request.form['jenis_tujuan']
        nama_tujuan = request.form['nama_tujuan']
        alamat_tujuan = request.form.get('alamat_tujuan', '')
        penerima = request.form.get('penerima', '')
        catatan = request.form.get('catatan', '')
        
        try:
            # Cek stok cukup
            cur.execute("SELECT jumlah FROM stok WHERE bahan_id = %s", [bahan_id])
            stok = cur.fetchone()
            
            if not stok or float(stok['jumlah']) < float(jumlah):
                flash('Stok tidak mencukupi!', 'danger')
            else:
                # Simpan pengeluaran
                cur.execute("""
                    INSERT INTO pengeluaran 
                    (no_pengeluaran, tanggal, bahan_id, jumlah, satuan_id, tujuan,
                     jenis_tujuan, nama_tujuan, alamat_tujuan, penerima, catatan, status)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, 'dikirim')
                """, (no_pengeluaran, tanggal, bahan_id, jumlah, satuan_id, tujuan,
                      jenis_tujuan, nama_tujuan, alamat_tujuan, penerima, catatan))
                
                # Update stok (kurangi)
                cur.execute("""
                    UPDATE stok 
                    SET jumlah = jumlah - %s,
                        last_update = CURRENT_TIMESTAMP
                    WHERE bahan_id = %s
                """, (jumlah, bahan_id))
                
                mysql.connection.commit()
                flash('Pengeluaran berhasil dicatat dan stok diperbarui!', 'success')
                return redirect(url_for('pengeluaran'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    # Ambil data untuk dropdown
    cur.execute("SELECT b.*, COALESCE(s.jumlah, 0) as stok_sekarang FROM bahan b LEFT JOIN stok s ON b.id = s.bahan_id WHERE b.status = 'aktif' ORDER BY b.nama_bahan")
    bahan_list = cur.fetchall()
    
    # Konversi Decimal ke float untuk stok
    for item in bahan_list:
        if isinstance(item.get('stok_sekarang'), Decimal):
            item['stok_sekarang'] = float(item['stok_sekarang'])
    
    cur.execute("SELECT * FROM satuan ORDER BY nama_satuan")
    satuan_list = cur.fetchall()
    
    # Generate nomor pengeluaran otomatis
    today = datetime.now().strftime('%Y%m%d')
    cur.execute("SELECT COUNT(*) as count FROM pengeluaran WHERE DATE(created_at) = CURDATE()")
    count = cur.fetchone()['count']
    no_pengeluaran_otomatis = f"KLR-{today}-{count+1:03d}"
    
    cur.close()
    
    return render_template('tambah_pengeluaran.html', 
                         bahan_list=bahan_list,
                         satuan_list=satuan_list,
                         no_pengeluaran_otomatis=no_pengeluaran_otomatis)

# Monitoring Kualitas
@app.route('/monitoring')
def monitoring():
    if not is_logged_in():
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Filter
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    bahan_id = request.args.get('bahan_id', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT mk.*, b.nama_bahan, b.kode_bahan
        FROM monitoring_kualitas mk
        JOIN bahan b ON mk.bahan_id = b.id
        WHERE 1=1
    """
    params = []
    
    if start_date:
        query += " AND mk.tanggal_check >= %s"
        params.append(start_date)
    
    if end_date:
        query += " AND mk.tanggal_check <= %s"
        params.append(end_date)
    
    if bahan_id:
        query += " AND mk.bahan_id = %s"
        params.append(bahan_id)
    
    query += " ORDER BY mk.tanggal_check DESC"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Ambil list bahan untuk filter
    cur.execute("SELECT * FROM bahan WHERE status = 'aktif' ORDER BY nama_bahan")
    bahan_list = cur.fetchall()
    
    cur.close()
    
    return render_template('monitoring.html', 
                         monitoring=data,
                         start_date=start_date,
                         end_date=end_date,
                         bahan_id=bahan_id,
                         bahan_list=bahan_list)

# Tambah Monitoring Kualitas
@app.route('/tambah-monitoring', methods=['GET', 'POST'])
def tambah_monitoring():
    if not check_role(['admin', 'gudang']):
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    cur = mysql.connection.cursor()
    
    if request.method == 'POST':
        bahan_id = request.form['bahan_id']
        tanggal_check = request.form['tanggal_check']
        suhu_gudang = float(request.form.get('suhu_gudang', 0)) if request.form.get('suhu_gudang') else None
        kelembaban_gudang = float(request.form.get('kelembaban_gudang', 0)) if request.form.get('kelembaban_gudang') else None
        kondisi_fisik = request.form.get('kondisi_fisik', 'baik')
        kondisi_kemasan = request.form.get('kondisi_kemasan', 'utuh')
        status_kadaluarsa = request.form.get('status_kadaluarsa', 'aman')
        petugas = request.form.get('petugas', '')
        catatan = request.form.get('catatan', '')
        
        try:
            cur.execute("""
                INSERT INTO monitoring_kualitas 
                (bahan_id, tanggal_check, suhu_gudang, kelembaban_gudang,
                 kondisi_fisik, kondisi_kemasan, status_kadaluarsa, petugas, catatan)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (bahan_id, tanggal_check, suhu_gudang, kelembaban_gudang,
                  kondisi_fisik, kondisi_kemasan, status_kadaluarsa, petugas, catatan))
            
            mysql.connection.commit()
            flash('Monitoring kualitas berhasil dicatat!', 'success')
            return redirect(url_for('monitoring'))
        except Exception as e:
            mysql.connection.rollback()
            flash(f'Error: {str(e)}', 'danger')
    
    # Ambil data untuk dropdown
    cur.execute("SELECT * FROM bahan WHERE status = 'aktif' ORDER BY nama_bahan")
    bahan_list = cur.fetchall()
    
    cur.close()
    
    return render_template('tambah_monitoring.html', 
                         bahan_list=bahan_list)

# Laporan Stok
@app.route('/laporan-stok')
def laporan_stok():
    if not is_logged_in():
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Filter
    kategori_id = request.args.get('kategori_id', '')
    stok_minimum = request.args.get('stok_minimum', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT b.*, kb.nama_kategori, s.nama_satuan, COALESCE(st.jumlah, 0) as stok_sekarang,
               CASE 
                   WHEN COALESCE(st.jumlah, 0) <= b.stok_minimum THEN 'kritis'
                   WHEN COALESCE(st.jumlah, 0) <= b.stok_minimum * 1.5 THEN 'rendah'
                   ELSE 'aman'
               END as status_stok
        FROM bahan b
        LEFT JOIN kategori_bahan kb ON b.kategori_id = kb.id
        LEFT JOIN satuan s ON b.satuan_id = s.id
        LEFT JOIN stok st ON b.id = st.bahan_id
        WHERE b.status = 'aktif'
    """
    params = []
    
    if kategori_id:
        query += " AND b.kategori_id = %s"
        params.append(kategori_id)
    
    if stok_minimum == 'ya':
        query += " AND COALESCE(st.jumlah, 0) <= b.stok_minimum"
    
    query += " ORDER BY status_stok, b.nama_bahan"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Konversi Decimal ke float untuk field tertentu
    for item in data:
        for key in ['stok_sekarang', 'stok_minimum', 'stok_maksimum']:
            if item.get(key) is not None and isinstance(item[key], Decimal):
                item[key] = float(item[key])
    
    # Hitung total nilai stok
    total_nilai = 0
    for item in data:
        # Ambil harga rata-rata dari penerimaan
        cur.execute("""
            SELECT AVG(harga_satuan) as harga_rata 
            FROM penerimaan 
            WHERE bahan_id = %s AND harga_satuan > 0
        """, [item['id']])
        harga = cur.fetchone()
        if harga and harga['harga_rata']:
            total_nilai += float(harga['harga_rata']) * float(item['stok_sekarang'] if item['stok_sekarang'] else 0)
    
    # Ambil list kategori untuk filter
    cur.execute("SELECT * FROM kategori_bahan ORDER BY nama_kategori")
    kategori_list = cur.fetchall()
    
    cur.close()
    
    return render_template('laporan_stok.html', 
                         stok=data,
                         kategori_id=kategori_id,
                         stok_minimum=stok_minimum,
                         kategori_list=kategori_list,
                         total_nilai=total_nilai)

# Laporan Distribusi
@app.route('/laporan-distribusi')
def laporan_distribusi():
    if not is_logged_in():
        flash('Akses ditolak!', 'danger')
        return redirect(url_for('dashboard'))
    
    # Filter
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    jenis_tujuan = request.args.get('jenis_tujuan', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT p.*, b.nama_bahan, b.kode_bahan, s.nama_satuan,
               CASE p.jenis_tujuan
                   WHEN 'sekolah' THEN 'Sekolah'
                   WHEN 'posyandu' THEN 'Posyandu'
                   WHEN 'puskesmas' THEN 'Puskesmas'
                   WHEN 'rumah_sakit' THEN 'Rumah Sakit'
                   ELSE 'Lainnya'
               END as nama_jenis_tujuan
        FROM pengeluaran p
        JOIN bahan b ON p.bahan_id = b.id
        JOIN satuan s ON p.satuan_id = s.id
        WHERE p.status != 'draft'
    """
    params = []
    
    if start_date:
        query += " AND p.tanggal >= %s"
        params.append(start_date)
    
    if end_date:
        query += " AND p.tanggal <= %s"
        params.append(end_date)
    
    if jenis_tujuan:
        query += " AND p.jenis_tujuan = %s"
        params.append(jenis_tujuan)
    
    query += " ORDER BY p.tanggal DESC"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Konversi Decimal ke float untuk field tertentu
    for item in data:
        for key in ['jumlah']:
            if item.get(key) is not None and isinstance(item[key], Decimal):
                item[key] = float(item[key])
    
    # Hitung total distribusi
    total_jumlah = sum(float(item['jumlah']) for item in data) if data else 0
    
    # Hitung distribusi per tujuan
    cur.execute("""
        SELECT jenis_tujuan, COUNT(*) as jumlah_transaksi, SUM(jumlah) as total_jumlah
        FROM pengeluaran
        WHERE status != 'draft'
        GROUP BY jenis_tujuan
    """)
    distribusi_per_tujuan = cur.fetchall()
    
    # Konversi Decimal ke float untuk distribusi_per_tujuan
    for item in distribusi_per_tujuan:
        if isinstance(item.get('total_jumlah'), Decimal):
            item['total_jumlah'] = float(item['total_jumlah'])
    
    cur.close()
    
    return render_template('laporan_distribusi.html', 
                         distribusi=data,
                         start_date=start_date,
                         end_date=end_date,
                         jenis_tujuan=jenis_tujuan,
                         total_jumlah=total_jumlah,
                         distribusi_per_tujuan=distribusi_per_tujuan)

# Export Laporan Stok ke PDF
@app.route('/export-stok-pdf')
def export_stok_pdf():
    if not is_logged_in():
        return redirect(url_for('login'))
    
    # Filter
    kategori_id = request.args.get('kategori_id', '')
    stok_minimum = request.args.get('stok_minimum', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT b.kode_bahan, b.nama_bahan, kb.nama_kategori, s.nama_satuan, 
               COALESCE(st.jumlah, 0) as stok_sekarang, b.stok_minimum,
               CASE 
                   WHEN COALESCE(st.jumlah, 0) <= b.stok_minimum THEN 'KRITIS'
                   WHEN COALESCE(st.jumlah, 0) <= b.stok_minimum * 1.5 THEN 'RENDAH'
                   ELSE 'AMAN'
               END as status_stok
        FROM bahan b
        LEFT JOIN kategori_bahan kb ON b.kategori_id = kb.id
        LEFT JOIN satuan s ON b.satuan_id = s.id
        LEFT JOIN stok st ON b.id = st.bahan_id
        WHERE b.status = 'aktif'
    """
    params = []
    
    if kategori_id:
        query += " AND b.kategori_id = %s"
        params.append(kategori_id)
    
    if stok_minimum == 'ya':
        query += " AND COALESCE(st.jumlah, 0) <= b.stok_minimum"
    
    query += " ORDER BY status_stok, b.nama_bahan"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Konversi Decimal ke float untuk field tertentu
    for item in data:
        for key in ['stok_sekarang', 'stok_minimum']:
            if item.get(key) is not None and isinstance(item[key], Decimal):
                item[key] = float(item[key])
    
    # Buat buffer untuk PDF
    buffer = io.BytesIO()
    
    # Buat dokumen PDF
    doc = SimpleDocTemplate(buffer, pagesize=landscape(A4))
    
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Judul
    title = Paragraph("LAPORAN STOK BAHAN SPPG - PROGRAM MBG", title_style)
    elements.append(title)
    
    # Info filter
    filter_text = "Tanggal: " + datetime.now().strftime("%d-%m-%Y")
    if kategori_id:
        cur.execute("SELECT nama_kategori FROM kategori_bahan WHERE id = %s", [kategori_id])
        kategori = cur.fetchone()
        if kategori:
            filter_text += f" | Kategori: {kategori['nama_kategori']}"
    
    filter_style = ParagraphStyle(
        'FilterStyle',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=15,
        alignment=TA_CENTER
    )
    filter_para = Paragraph(filter_text, filter_style)
    elements.append(filter_para)
    
    elements.append(Spacer(1, 20))
    
    # Siapkan data tabel
    table_data = []
    
    # Header tabel
    headers = ['No', 'Kode', 'Nama Bahan', 'Kategori', 'Stok', 'Satuan', 'Stok Min', 'Status']
    table_data.append(headers)
    
    # Isi data
    for i, item in enumerate(data, 1):
        row = [
            str(i),
            item['kode_bahan'],
            item['nama_bahan'],
            item['nama_kategori'],
            f"{item['stok_sekarang']:,.0f}",
            item['nama_satuan'],
            f"{item['stok_minimum']:,.0f}",
            item['status_stok']
        ]
        table_data.append(row)
    
    # Buat tabel
    table = Table(table_data, repeatRows=1)
    
    # Style tabel
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
    ])
    
    # Warna status
    for i, item in enumerate(data, 1):
        if item['status_stok'] == 'KRITIS':
            style.add('BACKGROUND', (7, i), (7, i), colors.red)
            style.add('TEXTCOLOR', (7, i), (7, i), colors.white)
        elif item['status_stok'] == 'RENDAH':
            style.add('BACKGROUND', (7, i), (7, i), colors.yellow)
        else:
            style.add('BACKGROUND', (7, i), (7, i), colors.green)
            style.add('TEXTCOLOR', (7, i), (7, i), colors.white)
    
    table.setStyle(style)
    elements.append(table)
    
    # Footer
    elements.append(Spacer(1, 20))
    footer_style = ParagraphStyle(
        'FooterStyle',
        parent=styles['Normal'],
        fontSize=9,
        alignment=TA_CENTER
    )
    
    current_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    footer_text = f"Dicetak pada: {current_time} | SPPG - Sistem Inventaris Gudang MBG"
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    
    # Buat nama file
    filename = f"laporan_stok_sppg_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/pdf'
    )

# Export Laporan Distribusi ke Excel
@app.route('/export-distribusi-excel')
def export_distribusi_excel():
    if not is_logged_in():
        return redirect(url_for('login'))
    
    # Filter
    start_date = request.args.get('start_date', '')
    end_date = request.args.get('end_date', '')
    jenis_tujuan = request.args.get('jenis_tujuan', '')
    
    cur = mysql.connection.cursor()
    
    query = """
        SELECT p.tanggal, p.no_pengeluaran, b.kode_bahan, b.nama_bahan,
               p.jumlah, s.nama_satuan, p.jenis_tujuan, p.nama_tujuan,
               p.alamat_tujuan, p.penerima, p.catatan, p.status
        FROM pengeluaran p
        JOIN bahan b ON p.bahan_id = b.id
        JOIN satuan s ON p.satuan_id = s.id
        WHERE p.status != 'draft'
    """
    params = []
    
    if start_date:
        query += " AND p.tanggal >= %s"
        params.append(start_date)
    
    if end_date:
        query += " AND p.tanggal <= %s"
        params.append(end_date)
    
    if jenis_tujuan:
        query += " AND p.jenis_tujuan = %s"
        params.append(jenis_tujuan)
    
    query += " ORDER BY p.tanggal DESC"
    
    cur.execute(query, params)
    data = cur.fetchall()
    
    # Konversi Decimal ke float untuk field tertentu
    for item in data:
        for key in ['jumlah']:
            if item.get(key) is not None and isinstance(item[key], Decimal):
                item[key] = float(item[key])
    
    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Laporan Distribusi"
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Border
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Judul
    ws.merge_cells('A1:L1')
    ws['A1'] = "LAPORAN DISTRIBUSI BAHAN SPPG - PROGRAM MBG"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Filter info
    filter_text = ""
    if start_date:
        filter_text += f"Periode: {start_date}"
        if end_date:
            filter_text += f" s/d {end_date}"
    
    ws.merge_cells('A2:L2')
    ws['A2'] = filter_text if filter_text else "Semua Data"
    ws['A2'].font = Font(italic=True, size=10)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Header tabel
    headers = ['No', 'Tanggal', 'No Pengeluaran', 'Kode Bahan', 'Nama Bahan',
               'Jumlah', 'Satuan', 'Jenis Tujuan', 'Nama Tujuan', 'Penerima', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Isi data
    for i, item in enumerate(data, 1):
        row_num = i + 4
        
        # Map status
        status_map = {
            'draft': 'Draft',
            'dikirim': 'Dikirim',
            'diterima': 'Diterima'
        }
        
        # Map jenis tujuan
        jenis_map = {
            'sekolah': 'Sekolah',
            'posyandu': 'Posyandu',
            'puskesmas': 'Puskesmas',
            'rumah_sakit': 'Rumah Sakit',
            'lainnya': 'Lainnya'
        }
        
        ws.cell(row=row_num, column=1, value=i).border = thin_border
        ws.cell(row=row_num, column=2, value=str(item['tanggal'])).border = thin_border
        ws.cell(row=row_num, column=3, value=item['no_pengeluaran']).border = thin_border
        ws.cell(row=row_num, column=4, value=item['kode_bahan']).border = thin_border
        ws.cell(row=row_num, column=5, value=item['nama_bahan']).border = thin_border
        ws.cell(row=row_num, column=6, value=item['jumlah']).border = thin_border
        ws.cell(row=row_num, column=7, value=item['nama_satuan']).border = thin_border
        ws.cell(row=row_num, column=8, value=jenis_map.get(item['jenis_tujuan'], item['jenis_tujuan'])).border = thin_border
        ws.cell(row=row_num, column=9, value=item['nama_tujuan']).border = thin_border
        ws.cell(row=row_num, column=10, value=item['penerima']).border = thin_border
        ws.cell(row=row_num, column=11, value=status_map.get(item['status'], item['status'])).border = thin_border
    
    # Summary
    summary_row = len(data) + 6
    ws.merge_cells(f'A{summary_row}:E{summary_row}')
    ws.cell(row=summary_row, column=1, value="TOTAL DISTRIBUSI:").font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=sum(item['jumlah'] for item in data)).font = Font(bold=True)
    
    # Set lebar kolom
    column_widths = [5, 12, 18, 12, 25, 10, 8, 12, 20, 15, 10]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Footer
    current_time = datetime.now().strftime("%d-%m-%Y %H:%M:%S")
    footer_row = len(data) + 8
    ws.cell(row=footer_row, column=1, value=f"Dicetak pada: {current_time}")
    ws.cell(row=footer_row, column=1).font = Font(size=9, italic=True)
    
    # Simpan ke buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    # Buat nama file
    filename = f"laporan_distribusi_sppg_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    cur.close()
    
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)